import mimetypes
import openpyxl
import os

from django.conf import settings
from django.contrib.auth.decorators import user_passes_test
from django.http.response import HttpResponse
from django.shortcuts import render, redirect
from django.urls import reverse

from .forms import UploadFileForm
from pipes_shop.models import Manual, Registry, Note, Pipe


def fill_first_row(worksheet, notes):
    index = 2
    for note in notes.order_by('name'):
        worksheet.cell(row=1, column=index).value = note.name
        index += 1


def fill_row(worksheet, registry, manual, pipes_line, row_index):
    index = 2
    worksheet.cell(row=row_index, column=1).value = registry.name
    for note in manual.notes.order_by('name'):
        for pipe in pipes_line:
            if pipe.note == note:
                worksheet.cell(row=row_index, column=index).value = pipe.price
        index += 1


def database_export_to_excel():
    work_book = openpyxl.Workbook()
    sheets = {}

    for manual in Manual.objects.all():
        sheets[manual.name] = 2
        work_book.create_sheet(manual.name)
        fill_first_row(work_book[manual.name], manual.notes)

        work_sheet = work_book[manual.name]
        pipes = Pipe.objects.all().filter(manual=manual)
        for registry in manual.registries.all():
            single_line = pipes.filter(registry=registry)
            fill_row(work_sheet, registry, manual, single_line, sheets[manual.name])
            sheets[manual.name] += 1

    del work_book['Sheet']
    work_book.save('database_export.xlsx')


@user_passes_test(lambda u: u.is_admin)
def generate_xlsx(request):
    database_export_to_excel()
    return redirect("cart:home")


@user_passes_test(lambda u: u.is_admin)
def generate_xlsx_page(request):
    from orders.nordigen import make_paid_bank_transfers
    make_paid_bank_transfers()
    return render(request, 'creating-xlsx.html', {
        'action_url': reverse("files:download")
    })


def handle_uploaded_file(f):
    wb = openpyxl.load_workbook(f)
    for sheet_index in range(len(wb.worksheets)):
        ws = wb.worksheets[sheet_index]

        if ws.max_row < 2:
            continue

        current_manual = Manual.objects.create(name=ws.title)

        notes_dict = dict()

        # generate notes
        for index in range(2, ws.max_column):
            note_name = ws.cell(row=1, column=index).value
            note, _ = Note.objects.get_or_create(name=note_name)
            current_manual.notes.add(note)
            notes_dict[index] = note

        # generate registries
        for row_index in range(2, ws.max_row):
            registry_name = ws.cell(row=row_index, column=1).value
            registry, _ = Registry.objects.get_or_create(name=registry_name)
            current_manual.registries.add(registry)

            # generate pipes
            for index in range(2, ws.max_column):
                pipe_price = ws.cell(row=row_index, column=index).value
                current_note = notes_dict[index]

                if pipe_price is None or pipe_price == "":
                    continue

                Pipe.objects.create(name=str(current_manual.name + "-" + registry_name + "-" + current_note.name),
                                    registry=registry,
                                    note=current_note,
                                    manual=current_manual,
                                    price=int(pipe_price)
                                    )

        print(ws.title, ws.max_column, ws.max_row)


@user_passes_test(lambda u: u.is_admin)
def upload_file(request):
    if request.method == 'POST':
        form = UploadFileForm(request.POST, request.FILES)
        if form.is_valid():
            handle_uploaded_file(request.FILES['file'])
    else:
        form = UploadFileForm()
    return render(request, 'upload.html', {'form': form})


@user_passes_test(lambda u: u.is_admin)
def download_file(request):
    database_export_to_excel()
    base_dir = settings.BASE_DIR
    filename = 'database_export.xlsx'
    filepath = os.path.join(base_dir, filename)
    path = open(filepath, 'rb')
    mime_type, _ = mimetypes.guess_type(filepath)
    response = HttpResponse(path, content_type=mime_type)
    response['Content-Disposition'] = "attachment; filename=%s" % filename
    return response
